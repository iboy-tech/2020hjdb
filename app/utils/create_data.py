# -*- coding:UTF-8 -*-
# !/usr/bin/python
"""
@File    : create_data.py
@Time    : 2020/1/29 13:57
@Author  : iBoy
@Email   : iboy@iboy.tech
@Description : 
@Software: PyCharm
"""
from openpyxl import load_workbook

from app.models.user_model import User
from app import db

wb = load_workbook('total.xlsx')
sheet = wb['Sheet1']


# list=[]
def create_test_data():
    print('被调用了')
    for i in range(1, sheet.max_row):
        academy = sheet["A{}".format(i)].value
        major = sheet["B{}".format(i)].value
        class_name = sheet["C{}".format(i)].value
        real_name = sheet["D{}".format(i)].value
        username = sheet["E{}".format(i)].value
        gender = sheet["F{}".format(i)].value
        qq = sheet["G{}".format(i)].value
        user = User(username=username, password='123456', real_name=real_name, academy=academy, class_name=class_name,
                    major=major, qq=qq, gender=gender)
        print(user)
        db.session.add(user)
        try:
            db.session.commit()
        except:
            db.session.rollback()


#     # list.append(user)
# db.session.add_all(list)
# db.session.commit()
if __name__ == 'page':
    create_test_data()
