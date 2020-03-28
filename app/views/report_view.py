# -*- coding:UTF-8 -*-
# !/usr/bin/python
"""
@File    : report_view.py
@Time    : 2020/2/27 14:22
@Author  : iBoy
@Email   : iboy@iboy.tech
@Description : 
@Software: PyCharm
"""
import os
from datetime import datetime, timedelta
from uuid import uuid4

from flask import render_template, request
from flask_login import login_required, current_user
from openpyxl import load_workbook
from sqlalchemy import and_

from app import db, Report, User
from app.decorators import wechat_required, admin_required
from app.models.category_model import Category
from app.models.lostfound_model import LostFound
from app.page import report
from app.utils import restful


@report.route('/', methods=['POST', 'GET'], strict_slashes=False)
@login_required
@wechat_required
@admin_required
def get_report():
    return render_template('report.html')


@report.route('/getall', methods=['POST'], strict_slashes=False)
@login_required
@wechat_required
@admin_required
def get_file():
    list = db.session.query(Report).order_by(Report.create_time.desc()).all()
    mylist = []
    if list:
        for file in list:
            # print(type(file))
            # print(file.file_name)
            u = User.query.get(file.user_id)
            if u:
                name = u.real_name
            else:
                name = "未知"
            dict = {
                "id": file.id,
                "filename": file.file_name + ".xlsx",
                "create_time": file.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                "creater": name
            }
            mylist.append(dict)
    data = {"list": mylist}
    # print(data)
    return restful.success(data=data)


@report.route('/add', methods=['POST'], strict_slashes=False)
@login_required
@wechat_required
@admin_required
def create_report():
    req = request.json
    # print('生成数据的条件', req)
    if req.get('type') == '1':  # 失物登记表
        get_lost_record(req.get("start"), req.get("end"))
    elif req.get('type') == '2':  # 失物统计表
        get_lost_summary(req.get("start"), req.get("end"))
    else:
        return restful.params_error(msg="报表类型错误")
    return restful.success(msg="导出成功")


@report.route('/delete', methods=['POST'], strict_slashes=False)
@login_required
@wechat_required
@admin_required
def delete_report():
    id = request.json
    # print('要删除的文件', id, type(id))
    file = Report.query.filter_by(id=id).first()
    if file:
        filename = os.path.join(os.getenv('PATH_OF_REPORT'), id + '.xlsx')
        try:
            db.session.delete(file)
            db.session.commit()
            os.remove(filename)
        except Exception as e:
            db.session.rollback()
    return restful.success(msg="删除成功")


# 失物统计表（归还统计,失物统计）
def get_lost_record(start_time, end_time):
    """
    (归还统计)
    类型	失物数量	找回数量	拾取数量	归还数量
    (失物统计)
    失主/拾主	物品	拾取/丢失地点	所属学院
    :return:
    """
    start_time = datetime.strptime(start_time, '%Y-%m-%d')
    end_time = datetime.strptime(end_time, '%Y-%m-%d')
    found_list = db.session.query(LostFound).filter(
        and_(LostFound.create_time.between(start_time, end_time + timedelta(days=1)),
             LostFound.kind == 1)) \
        .order_by(LostFound.create_time).all()
    # print('我是查询的报表', found_list)
    lost_xlsx = load_workbook(os.path.join(os.getenv('PATH_OF_TEMP'), '失物登记表.xlsx'))
    sheet_names = lost_xlsx.sheetnames
    """
    丢失物品
    失主	失物	失物详情	遗失地点	遗失时间	失主联系方式
    """
    if found_list:
        star_row = 4
        for f in found_list:
            ws = lost_xlsx.get_sheet_by_name(sheet_names[0])
            mylist = [f.post_category.name, f.about, f.location, f.create_time.strftime('%Y-%m-%d'), f.post_user.qq]
            # ws.append(mylist)
            for i in range(1, 6):
                ws.cell(row=star_row, column=i, value=mylist[i - 1])
            star_row += 1

    found_list = db.session.query(LostFound).filter(
        and_(LostFound.create_time.between(start_time, end_time + timedelta(days=1)),
             LostFound.kind == 0)) \
        .order_by(LostFound.create_time).all()
    # print('我是查询的报表', found_list)
    """
    失主	失物	失物详情	遗失地点	遗失时间	失主联系方式
    """
    if found_list:
        star_row = 4
        for f in found_list:
            ws = lost_xlsx.get_sheet_by_name(sheet_names[1])
            mylist = [f.post_user.real_name, f.post_category.name, f.about, f.location,
                      f.create_time.strftime('%Y-%m-%d'), f.post_user.qq]
            # print(f.post_user.real_name, f.post_category.name,f.post_user.qq, f.about , f.location, f.create_time.strftime('%Y-%m-%d'),f.post_user.qq)
            for i in range(1, 7):
                ws.cell(row=star_row, column=i, value=mylist[i - 1])
            star_row += 1
    fid = uuid4().hex
    filename = fid + '.xlsx'
    xlsx_res = os.path.join(os.getenv('PATH_OF_REPORT'), filename)
    lost_xlsx.save(xlsx_res)
    try:
        new_report = Report(id=fid,
                            file_name=start_time.strftime('%Y-%m-%d') + '~' + end_time.strftime('%Y-%m-%d') + ' 失物登记表',
                            user_id=current_user.id)
        db.session.add(new_report)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        os.remove(xlsx_res)
        return restful.success(success=False, msg=str(e))
    pass


# 失物登记表（拾物采集表，失物上报表）
def get_lost_summary(start_time, end_time):
    """
    拾物采集kind=1
    """
    start_time = datetime.strptime(start_time, '%Y-%m-%d')
    end_time = datetime.strptime(end_time, '%Y-%m-%d')
    # found_list = db.session.query(LostFound).filter(
    #     LostFound.create_time.between(start_time, end_time + timedelta(days=1)),
    #     LostFound.kind == 1) \
    #     .order_by(LostFound.create_time).all()

    """
    (拾物采集表)
    物品	详情	拾取地点	拾取时间	领取方式
    (失物上报表)
    失主	失物	失物详情	遗失地点	遗失时间	失主联系方式
    :return:
    """
    """
    (归还统计)
    类型	失物数量	找回数量	拾取数量	归还数量
    """
    record_xlsx = load_workbook(os.path.join(os.getenv('PATH_OF_TEMP'), '失物统计表.xlsx'))
    sheet_names = record_xlsx.sheetnames
    categories = db.session.query(Category).all()
    star_row = 4
    for c in categories:
        mylist = []
        mylist.append(c.name)
        ws = record_xlsx.get_sheet_by_name(sheet_names[0])
        found_list = db.session.query(LostFound).filter(
            and_(LostFound.category_id == c.id, LostFound.kind == 0, LostFound.status == 0,
                 LostFound.create_time.between(start_time, end_time + timedelta(days=1)))).count()
        # print(start_time, end_time, '失物数量查询', found_list)
        mylist.append(found_list)
        found_list = db.session.query(LostFound).filter(
            and_(LostFound.category_id == c.id, LostFound.kind == 0, LostFound.status == 1,
                 LostFound.create_time.between(start_time, end_time + timedelta(days=1)))).count()
        # print(start_time, end_time, '找回数量查询', found_list)
        mylist.append(found_list)
        found_list = db.session.query(LostFound).filter(
            and_(LostFound.category_id == c.id, LostFound.kind == 1, LostFound.status == 0,
                 LostFound.create_time.between(start_time, end_time + timedelta(days=1)))).count()
        # print(start_time, end_time, '拾取数量查询', found_list)
        mylist.append(found_list)
        found_list = db.session.query(LostFound).filter(
            and_(LostFound.category_id == c.id, LostFound.kind == 1, LostFound.status == 1,
                 LostFound.create_time.between(start_time, end_time + timedelta(days=1)))).count()
        # print(start_time, end_time, '归还数量查询', found_list)
        mylist.append(found_list)
        for i in range(1, 6):
            ws.cell(row=star_row, column=i, value=mylist[i - 1])
        star_row += 1
    """
    (失物统计)
    失主/拾主	物品	拾取/丢失地点	所属学院
    """
    found_list = db.session.query(LostFound).filter(
        and_(LostFound.create_time.between(start_time, end_time + timedelta(days=1)))) \
        .order_by(LostFound.create_time).all()
    # print('我是查询的报表', found_list)
    star_row = 4
    if found_list:
        for f in found_list:
            ws = record_xlsx.get_sheet_by_name(sheet_names[1])
            mylist = [f.post_user.real_name, f.post_category.name, f.location, f.post_user.academy]
            for i in range(1, 5):
                ws.cell(row=star_row, column=i, value=mylist[i - 1])
            star_row += 1
    fid = uuid4().hex
    filename = fid + '.xlsx'
    xlsx_res = os.path.join(os.getenv('PATH_OF_REPORT'), filename)
    record_xlsx.save(xlsx_res)
    try:
        new_report = Report(id=fid,
                            file_name=start_time.strftime('%Y-%m-%d') + '~' + end_time.strftime('%Y-%m-%d') + ' 失物统计表',
                            user_id=current_user.id)
        db.session.add(new_report)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        os.remove(xlsx_res)
        return restful.success(success=False, msg=str(e))
